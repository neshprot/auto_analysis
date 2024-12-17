from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side


def make_exel_file(params, file_name, best_correlations, correlations):
    column_names = ['pKa', 'dih angle', 'fitness']
    correlation_column_names = ['increase pKa', 'decrease pKa', 'increase dih angle', 'dec dih angle']
    wb = Workbook()
    ws = wb.active
    letter_first = 65
    step = 8    # 4 - это отступ в один столбец между
    # Заголовки колонок
    for column_idx in range(len(params)):
        ws[f"{chr(letter_first)}1"] = f"Parameter {column_names[column_idx]}"
        ws[f"{chr(letter_first)}2"] = "mutations"
        ws[f"{chr(letter_first + 1)}2"] = "the first par"
        ws[f"{chr(letter_first + 2)}2"] = "the second par"

        ws[f"{chr(letter_first + 3)}2"] = correlation_column_names[0]
        ws[f"{chr(letter_first + 4)}2"] = correlation_column_names[1]
        ws[f"{chr(letter_first + 5)}2"] = correlation_column_names[2]
        ws[f"{chr(letter_first + 6)}2"] = correlation_column_names[3]
        letter_first += step
    colors = ["7cedac", "97e6ed", "dd76ef", "9f88ec"]
    for i, slices in enumerate(params):
        for j, muts in enumerate(slices):
            cell = ws.cell(row=j + 3, column=i*step + 1)
            cell.value = muts[0]
            cell = ws.cell(row=j + 3, column=i * step + 2)
            cell.value = round(muts[1][0], 2)
            cell = ws.cell(row=j + 3, column=i * step + 3)
            cell.value = round(muts[1][1], 2)
            for col_idx, cor in enumerate(best_correlations):
                if any([x[0] in muts[0] for x in cor]):
                    fill = PatternFill(start_color=colors[col_idx], end_color=colors[col_idx], fill_type="solid")
                    cell = ws.cell(row=j + 3, column=i * step + 4 + col_idx)
                    cell.fill = fill

    for idx, coord in zip([1, 2, 3], [f'Y2', 'Y20', 'AK2']):
        image_path = f'his_{idx}.png'
        img = Image(image_path)
        ws.add_image(img, coord)
        # Сохранение изменений
    row_start = len(params[0])
    new_step = 4
    column_names_for_corr = ('mut', 'correlation', 'p-value')
    for j, name_cor in zip((0, new_step), ['pKa cor', 'dih angle cor']):
        cell = ws.cell(row=6 + row_start, column=1 + j)
        cell.value = name_cor
        for i, name in enumerate(column_names_for_corr):
            cell = ws.cell(row=7 + row_start, column=i + 1 + j)
            cell.value = name
    decimal_places = 3
    for i, slices in enumerate(correlations):
        was_null = True
        for j, muts in enumerate(slices):
            cell = ws.cell(row=j + 8 + row_start, column=i * new_step + 1)
            cell.value = muts[0]
            cell = ws.cell(row=j + 8 + row_start, column=i * new_step + 2)
            cell.number_format = f"#,##0.{decimal_places * '0'}"
            cell.value = round(muts[1], decimal_places)
            cell = ws.cell(row=j + 8 + row_start, column=i * new_step + 3)
            cell.number_format = f"#,##0.{decimal_places * '0'}"
            cell.value = round(muts[2], decimal_places)
            if round(muts[2], 3) != 0 and was_null:
                was_null = False
                side = Side(border_style="medium", color='FF0000')
                border = Border(top=side, left=None, right=None, bottom=None)
                cell.border = border
                cell = ws.cell(row=j + 8 + row_start, column=i * new_step + 1)
                cell.border = border
                cell = ws.cell(row=j + 8 + row_start, column=i * new_step + 2)
                cell.border = border
    name_best_cor = ['Mutation increasing delta pKa', 'Mutation decreasing delta pKa',
                     'Mutation increasing dihedral angle', 'Mutation decreasing dihedral angle']
    slider = 0
    for cor, name in zip(best_correlations, name_best_cor):
        if cor:
            cell = ws.cell(row=6 + row_start, column=slider + 10)
            cell.value = name
            for idx, mut in enumerate(cor):
                cell = ws.cell(row=7 + row_start + idx, column=slider + 10)
                cell.value = mut[0]
            slider += 1

    wb.save(f"output_{file_name}.xlsx")
